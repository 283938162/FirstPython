# coding:utf-8
import requests
# import xlrd


import xlrd, xlwt

# openpyxl 用于读写2007版本的excel
import openpyxl

from os import path

from xlutils.copy import copy

from bs4 import BeautifulSoup
import os

from xlutils.copy import copy

'''
2007版以前的Excel（xls结尾的），需要使用xlrd读，xlwt写。  xl-xls wt-write rd-read   
2007版以后的Excel（xlsx结尾的），需要使用openpyxl来读写。

保存文件时必须要有路径参数：否则
TypeError: save() missing 1 required positional argument: 'filename_or_stream'

'''

url = 'http://www.gushiwen.org/'

path = 'data/gushici.xls'

values = []

# 给excel 加标头
value_head = ['标题', '朝代', '内容', '标签']
values.append(value_head)


def get_pages():
    testexist(path)
    html = requests.get(url)
    soup = BeautifulSoup(html.text, 'html.parser')
    pages = soup.find('div', class_="pages").find_all('a')
    page_max = pages[3].text
    # global num2
    # num2=len(page_max)+1
    for i in range(int(page_max) + 1):
        page_url = url + 'default_' + str(i) + '.aspx'

        html = requests.get(page_url)  # ,headers=header)

        soup = BeautifulSoup(html.text, 'html.parser')
        # print(soup)
        aa = soup.find_all('div', class_="sons", limit=10)
        # print(aa)
        # global num1
        # num1=len(aa)
        # global title

        for item in aa:
            title = item.b.text
            d_a = item.find_all('a')
            dynasty = d_a[1].text
            author = d_a[2].text
            content = item.find('div', class_="contson").text
            try:
                tag_ = item.find('div', class_="tag").a.text
                n = len(tag_)
                tag = ''
                for i in range(n):
                    tag = tag + str(tag_[i])
                # print(tag)
            except BaseException:
                tag = ''

            # 将 每一首诗词的每个字段放入一个list
            value = []
            value.append(title)
            value.append(dynasty)
            value.append(content)
            value.append(tag)

            print(value)
            # 将每一首诗词最为一个list元素 再放入一个list
            # values.append(value)

            # 逐条保存

            # 参数传递的形式 可以是任意自定义的  可以是多个字段 也可以将多个字段放入一个list（推荐)
            # add_excel_data_v(value)

            # 加标题头
            add_excel_data(title, dynasty, content, tag)


def testexist(path):
    if not os.path.exists(path):
        w = xlwt.Workbook()
        w.add_sheet('Sheet1')
        w.save(path)
    else:
        os.remove(path)
        w = xlwt.Workbook()
        w.add_sheet('Sheet1')
        w.save(path)


def add_excel_data_v(values):
    wb = xlrd.open_workbook(path)  # 用wlrd提供的方法读取一个excel文件

    # sheets = wb.sheet_names()
    # sheet = wb.sheet_by_name(sheets[0])

    rows = wb.sheets()[0].nrows  # 用wlrd提供的方法获得现在已有的行数

    new_wb = copy(wb)  # 用xlutils提供的copy方法将xlrd的对象转化为xlwt的对象

    new_sheet = new_wb.get_sheet(0)  # 用xlwt对象的方法获得要操作的sheet

    # values = ["1", "2", "3"]

    # for v in values:
    new_sheet.write(rows, 0, values[0])  # xlwt对象的写方法，参数分别是行、列、值

    new_sheet.write(rows, 1, values[1])
    new_sheet.write(rows, 2, values[2])
    new_sheet.write(rows, 3, values[3])

    new_wb.save(path)  # xlwt对象的保存方法，这时便覆盖掉了原来的excel
    print('数据追加成功!')


# 逐条追加 效率低下
def add_excel_data(title, dynasty, content, tag):
    wb = xlrd.open_workbook(path)  # 用wlrd提供的方法读取一个excel文件

    # sheets = wb.sheet_names()
    # sheet = wb.sheet_by_name(sheets[0])

    rows = wb.sheets()[0].nrows  # 用wlrd提供的方法获得现在已有的行数

    new_wb = copy(wb)  # 用xlutils提供的copy方法将xlrd的对象转化为xlwt的对象

    new_sheet = new_wb.get_sheet(0)  # 用xlwt对象的方法获得要操作的sheet  注意这个是（0） 不是[0]

    # new_sheet.write(rows, 0, times)  # xlwt对象的写方法，参数分别是行、列、值
    #
    # new_sheet.write(rows, 1, shuos)
    # print(rows)

    # 加说明头
    if rows == 0:
        new_sheet.write(0, 0, '标题')  # xlwt对象的写方法，参数分别是行、列、值
        new_sheet.write(0, 1, '朝代')
        new_sheet.write(0, 2, '内容')
        new_sheet.write(0, 3, '标题')

        new_sheet.write(rows + 1, 0, title)  # xlwt对象的写方法，参数分别是行、列、值

        new_sheet.write(rows + 1, 1, dynasty)
        new_sheet.write(rows + 1, 2, content)
        new_sheet.write(rows + 1, 3, tag)

    else:
        new_sheet.write(rows, 0, title)  # xlwt对象的写方法，参数分别是行、列、值

        new_sheet.write(rows, 1, dynasty)
        new_sheet.write(rows, 2, content)
        new_sheet.write(rows, 3, tag)

    new_wb.save(path)  # xlwt对象的保存方法，这时便覆盖掉了原来的excel
    print('数据追加成功!')


# 将所有的数据临时放到一个list的二维集合，抓取所有页面后 一次性写入
def save_values_toexcel(values, path, sheetname):
    wb = xlwt.Workbook()  # tab 自动没有（）
    sheet = wb.add_sheet(sheetname)
    rows = len(values)
    cols = len(values[0])

    for i in range(0, rows):
        for j in range(0, cols):
            sheet.write(i, j, values[i][j])
    wb.save(path)


if __name__ == '__main__':
    get_pages()
